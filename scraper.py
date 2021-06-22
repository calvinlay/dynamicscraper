from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
import requests

import time

options = webdriver.ChromeOptions() #Browser settings
options.add_argument('headless')
options.add_experimental_option('excludeSwitches', ['enable-logging'])

url = "https://yourcareer.gov.au/your-future-career/your-results/?includeExperienceAndQualifications=false&industries=J&interests=1%2C2%2C3%2C4%2C5%2C6%2COUTDOORS" #Page to scrape
driver = webdriver.Chrome(options=options, executable_path=r'C:\Users\Calvin.Nguyen\Desktop\yourcareer\chromedriver_win32\chromedriver.exe')
driver.get(url)
html = driver.page_source.encode('utf-8')
page_num = 0

while page_num < 12: #Number of times to select read more button
    driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//button[normalize-space()="Load more"]')))) #Search for button with "Load More"
    page_num += 1
    print("getting page number "+str(page_num))
    time.sleep(1)

html = driver.page_source.encode('utf-8') #Save result html

soup = BeautifulSoup(html, 'lxml') 
mydivs = soup.find_all("a", {"class": "c-careers-list__card-info-link blue-ex"}) #Find all divs with class name
wb = openpyxl.load_workbook(r'c:yourcareer.xlsx') #Open excel file
sheet = wb['Sheet1'] #Add sheet to read/write
it = 1
for res in mydivs:
    print(res.text)
    sheet.cell(row=it, column=1).value = res.text #Save text to excel document
    sheet.cell(row=it, column=2).value = res['href'] #Save Hyper link to excel document
    it = it + 1
            
wb.save('result.xlsx') # save excel document